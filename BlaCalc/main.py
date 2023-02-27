from charge import *
from cent_tb import *
from out_tb import *
from cent_pr import *
from out_pr import *
from sheet import *
# from openpyxl.styles import NamedStyle, Font, Border, Side#,='thin'

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl import Workbook


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def PCutTest():
    sh1 = sheet()

    b = 0.5
    d = 0.265
    print(f' cc of burden (b={b}m) is {cc_b(0.5)}')
    print(f' cc of C-C distance (d={d}m) is {cc_d(d)}')

    for d in range(5):
        print(f' cc of C-C distance (d={d/10}m) is {cc_d(d/10)}')

    param = {}
    param['A'] = 3.3
    param['H'] = 3.3
    # param['H2'] = 0 # for V-Cut, not used here
    # param['H3'] = 0 # for V-Cut, not used here
    param['Dia'] = 0.104
    param['n'] = 3
    param['IExp'] = mm_exp
    param['OExp'] = nf_exp

    pcut = PCut()
    print(pcut.name)
    pcut.set_param(param)
    pcut.calc()
    pcut.print_summary()

    stop = Stoping()
    stop.set_param(param)
    stop.calc(pcut.W)
    stop.print_summary()

    floor = Floor()
    floor.set_param(param)
    floor.calc(pcut.W)
    floor.print_summary()

    perim = Perim()
    perim.set_param(param)
    perim.calc(pcut.W)
    perim.print_summary()

    sh1.save(pcut)
    return pcut, stop, floor, perim

def PCutTestPR():
    sh1 = sheet()

    b = 0.5
    d = 0.265
    print(f' cc of burden (b={b}m) is {cc_b(0.5)}')
    print(f' cc of C-C distance (d={d}m) is {cc_d(d)}')

    for d in range(5):
        print(f' cc of C-C distance (d={d/10}m) is {cc_d(d/10)}')

    param = {}
    param['A'] = 3.3
    param['H'] = 3.3
    # param['H2'] = 0 # for V-Cut, not used here
    # param['H3'] = 0 # for V-Cut, not used here
    param['Dia'] = 0.104
    param['n'] = 3
    param['IExp'] = mm_exp
    param['OExp'] = nf_exp

    pcut = PCutPR()
    print(pcut.name)
    pcut.set_param(param)
    pcut.calc()
    pcut.print_summary()

    stop = StopingPR()
    stop.set_param(param)
    stop.calc(pcut.W)
    stop.print_summary()

    floor = FloorPR()
    floor.set_param(param)
    floor.calc(pcut.W)
    floor.print_summary()

    perim = PerimPR()
    perim.set_param(param)
    perim.calc(pcut.W)
    perim.print_summary()

    sh1.save(pcut)
    return pcut, stop, floor, perim


def VCutTest():
    sh1 = sheet()

    b = 0.5
    d = 0.265
    print(f' cc of burden (b={b}m) is {cc_b(0.5)}')
    print(f' cc of C-C distance (d={d}m) is {cc_d(d)}')

    for d in range(5):
        print(f' cc of C-C distance (d={d/10}m) is {cc_d(d/10)}')

    param = {}
    param['A'] = 1.1
    param['H'] = 0.66
    param['H2'] = 1.27
    param['H3'] = 1.21
    param['B2'] = 0.78
    param['IExp'] = em_exp
    param['OExp'] = nf_exp

    vcut = VCut()
    print(vcut.name)
    vcut.set_param(param)
    vcut.W = 0.7
    vcut.calc()
    vcut.print_summary()

    param['H'] = 1.1
    stop = Stoping()
    stop.set_param(param)
    stop.calc(vcut.W)
    stop.print_summary()

    floor = Floor()
    floor.set_param(param)
    floor.calc(vcut.W)
    floor.print_summary()

    perim = Perim()
    perim.set_param(param)
    perim.calc(vcut.W)
    perim.print_summary()

    sh1.save(vcut)


def rnd_test():
    for i in range(30):
        v = i/15
        print(v, rnd(v))


def xls_test():
    highlight = NamedStyle(name="highlight")
    bd = Side(style='thin', color="0000ff")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    thin_border = Border(
        left=Side(border_style='thin', color='00000000'),
        right=Side(border_style='thin', color='00000000'),
        top=Side(border_style='thin', color='00000000'),
        bottom=Side(border_style='thin', color='00000000')
    )

    write_wb = Workbook()
    write_wb.add_named_style(highlight)
    # write_wb.cell(row=3, column=2).border = thin_border

    # 이름이 있는 시트를 생성
    write_ws = write_wb.create_sheet('생성시트')

    # Sheet1에다 입력
    write_ws = write_wb.active
    write_ws['A1'] = '숫자'

    # 행 단위로 추가
    write_ws.append([1, 2, 3])

    # 셀 단위로 추가
    write_ws.cell(5, 5, '5행5열').style = highlight
    write_ws['B2:D5'].border = thin_border
    write_ws['B2'].style = highlight

    write_wb.save("d:/test.xlsx")


def form_bl(wb):
    wb = Workbook()
    ws = wb.active
    dotted = Side(border_style="dotted", color="000000")
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    double = Side(border_style="double", color="ff0000")

    tk_a_brd = Border(top=thick, left=thick, right=thick, bottom=thick)
    tk_tb_brd = Border(top=thick, left=thin, right=thin, bottom=thick)
    tk_tbl_brd = Border(top=thick, left=thick, right=thin, bottom=thick)
    tk_tbr_brd = Border(top=thick, left=thin, right=thick, bottom=thick)
    tk_a_brd = Border(top=thick, left=thick, right=thick, bottom=thick)
    tk_l_brd = Border(top=thin, left=thick, right=thin, bottom=thin)
    tk_r_brd = Border(top=thin, left=thin, right=thick, bottom=thin)
    tk_b_brd = Border(top=thin, left=thin, right=thin, bottom=thick)
    tk_br_brd = Border(top=thin, left=thin, right=thick, bottom=thick)
    tk_blr_brd = Border(top=thick, left=thick, right=thin, bottom=thick)
    tk_bl_brd = Border(top=thin, left=thick, right=thin, bottom=thick)
    thin_brd = Border(top=thin, left=thin, right=thin, bottom=thin)

    # pfill = PatternFill("solid", fgColor="DDDDDD")
    # gfill = GradientFill(stop=("000000", "FFFFFF"))
    hfont = Font(b=True, color="000000", size=16)
    tfont = Font(b=False, color="000000", size=10)
    al_c = Alignment(horizontal="center", vertical="center")
    al_l = Alignment(horizontal="left", vertical="center")
    al_r = Alignment(horizontal="right", vertical="center")

    width = [72, 72, 162, 6, 340, 72, 6, 340, 72]
    ws.column_dimensions['B'].width = width[0]/8.118
    ws.column_dimensions['C'].width = width[1]/8.118
    ws.column_dimensions['D'].width = width[2]/8.118
    ws.column_dimensions['E'].width = width[3]/8.118
    ws.column_dimensions['F'].width = width[4]/8.118
    ws.column_dimensions['G'].width = width[5]/8.118
    ws.column_dimensions['H'].width = width[6]/8.118
    ws.column_dimensions['I'].width = width[7]/8.118

    ws.row_dimensions[1].height = 57
    for i in range(2, 92):
        ws.row_dimensions[i].height = 24

    # style_range(ws, 'B2:F4', border=border, fill=pfill, font=font, alignment=al)
    # style_range(ws, 'B6:F8', border=border, fill=gfill, font=font, alignment=al)
    style_range(ws, 'B1:J1', border=tk_a_brd, font=hfont, alignment=al_c)
    style_range(ws, 'B2:C2', border=tk_tbl_brd, font=tfont, alignment=al_c)
    style_range(ws, 'D2:D2', border=tk_tb_brd, font=tfont, alignment=al_c)
    # style_range(ws, 'E2:E2', border=thin_brd, font=tfont, alignment=al)
    style_range(ws, 'F2:F2', border=tk_tb_brd, font=tfont, alignment=al_c)
    style_range(ws, 'G2:G2', border=tk_tb_brd, font=tfont, alignment=al_c)
    # style_range(ws, 'H2:H2', border=thin_brd, font=tfont, alignment=al)
    style_range(ws, 'I2:I2', border=tk_tb_brd, font=tfont, alignment=al_c)
    style_range(ws, 'J2:J2', border=tk_tbr_brd, font=tfont, alignment=al_c)

    str = 'H=4.4m'
    my_cell = ws['B1']
    my_cell.value = f'평행 심발공 패턴 설계 ({str})'
    ws['B2'] = '구분'
    ws['D2'] = '항목'
    ws['F2'] = 'Olloffson 식'
    ws['G2'] = '값'
    ws['I2'] = '현업 보고서 계산'
    ws['J2'] = '값'

    style_range(ws, 'B3:B43', border=tk_l_brd, font=tfont, alignment=al_c)
    style_range(ws, 'C3:C14', border=thin_brd, font=tfont, alignment=al_c)
    style_range(ws, 'C15:C22', border=thin_brd, font=tfont, alignment=al_c)
    style_range(ws, 'C23:C30', border=thin_brd, font=tfont, alignment=al_c)
    style_range(ws, 'C31:C43', border=thin_brd, font=tfont, alignment=al_c)

    ws['B3'] = '심발공'
    ws['C3'] = '심발공 1'
    ws['C15'] = '심발공 2'
    ws['C23'] = '심발공 3'
    ws['C31'] = '심발공 4'

    for i in range(3, 91):
        style_range(ws, f'D{i}:D{i}', border=thin_brd,
                    font=tfont, alignment=al_l)
        style_range(ws, f'F{i}:F{i}', border=thin_brd,
                    font=tfont, alignment=al_l)
        style_range(ws, f'G{i}:G{i}', border=thin_brd,
                    font=tfont, alignment=al_r)
        style_range(ws, f'I{i}:I{i}', border=thin_brd,
                    font=tfont, alignment=al_l)
        style_range(ws, f'J{i}:J{i}', border=tk_r_brd,
                    font=tfont, alignment=al_r)

    for i in range(91, 92):
        style_range(ws, f'D{i}:D{i}', border=tk_b_brd,
                    font=tfont, alignment=al_l)
        style_range(ws, f'F{i}:F{i}', border=tk_b_brd,
                    font=tfont, alignment=al_l)
        style_range(ws, f'G{i}:G{i}', border=tk_b_brd,
                    font=tfont, alignment=al_r)
        style_range(ws, f'I{i}:I{i}', border=tk_b_brd,
                    font=tfont, alignment=al_l)
        style_range(ws, f'J{i}:J{i}', border=tk_br_brd,
                    font=tfont, alignment=al_r)

    style_range(ws, 'B44:C44', border=tk_l_brd, font=tfont, alignment=al_c)
    style_range(ws, 'B45:C45', border=tk_l_brd, font=tfont, alignment=al_c)
    style_range(ws, 'B46:C46', border=tk_l_brd, font=tfont, alignment=al_c)

    ws['B44'] = '심발공 장약량'
    ws['B45'] = '암체부피'
    ws['B46'] = '비장약량'

    style_range(ws, 'B47:C60', border=tk_l_brd, font=tfont, alignment=al_c)
    style_range(ws, 'B61:C74', border=tk_l_brd, font=tfont, alignment=al_c)
    style_range(ws, 'B75:C88', border=tk_l_brd, font=tfont, alignment=al_c)

    ws['B47'] = '확대공'
    ws['B61'] = '바닥공'
    ws['B75'] = '외곽공'

    style_range(ws, 'B89:C89', border=tk_l_brd, font=tfont, alignment=al_c)
    style_range(ws, 'B90:C90', border=tk_l_brd, font=tfont, alignment=al_c)
    style_range(ws, 'B91:C91', border=tk_bl_brd, font=tfont, alignment=al_c)

    ws['B89'] = '총 장약량'
    ws['B90'] = '터널 굴진량'
    ws['B91'] = '총 비장약량'

    sim1 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '공당 장약밀도(Ic)',
        '공당장약량(Q)',
        '적용장약량(Qa)',
        '대구경 직경(φ)',
        '대구경 공수 (n)',
        '대구경 환산직경(φd)',
        '공공과의 거리(a)',
        '장약장 (L)',
        '사각형의 변(W1)']

    for i in range(3, 15):
        ws[f'D{i}'] = sim1[i-3]

    sim2 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '공당 장약밀도(Ic)',
        '공당장약량(Q)',
        '적용장약량(Qa)',
        '장약장 (L)',
        '사각형의 변(W2)']

    for i in range(15, 23):
        ws[f'D{i}'] = sim2[i-15]

    sim3 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '공당 장약밀도(Ic)',
        '공당장약량(Q)',
        '적용장약량(Qa)',
        '장약장 (L)',
        '사각형의 변(W3)']

    for i in range(23, 31):
        ws[f'D{i}'] = sim3[i-23]

    sim4 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '기저장 (hb)',
        '기저장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '공당장약량(Qc)',
        '적용 기저장약량(Qab)',
        '적용 주상장약량(Qac)',
        '총 장약량(Qtot)',
        '장약장 (L)',
        '사각형의 변(W4)']

    for i in range(31, 44):
        ws[f'D{i}'] = sim4[i-31]

    stoping = [
        '천공장 (H)',
        '저항선 (B)',
        '간격 (S)',
        '기저 장약장 (hb)',
        '전색장 (h0)',
        '기저 장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '주상장약량(Qc)',
        '적용기저장약량 (Qab)',
        '적용주상장약량 (Qac)',
        '총 공당장약량(Qtot)',
        '공수(n)',
        '장약장(L)']

    for i in range(47, 61):
        ws[f'D{i}'] = stoping[i-47]

    floor = [
        '천공장 (H)',
        '저항선 (B)',
        '간격 (S)',
        '기저 장약장 (hb)',
        '전색장 (h0)',
        '기저 장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '주상장약량(Qc)',
        '적용기저장약량 (Qab)',
        '적용주상장약량 (Qac)',
        '총 공당장약량(Qtot)',
        '공수(n)',
        '장약장(L)']

    for i in range(61, 75):
        ws[f'D{i}'] = floor[i-61]

    peri = [
        '천공장 (H)',
        '저항선 (B)',
        '간격 (S)',
        '기저 장약장 (hb)',
        '전색장 (h0)',
        '기저 장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '주상장약량(Qc)',
        '적용기저장약량 (Qab)',
        '적용주상장약량 (Qac)',
        '총 공당장약량(Qtot)',
        '공수(n)',
        '장약장(L)']

    for i in range(75, 89):
        ws[f'D{i}'] = peri[i-75]

    return wb


def xls_tb(wb):

    ws = wb.active

    sim1 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '공당 장약밀도(Ic)',
        '공당장약량(Q)',
        '적용장약량(Qa)',
        '대구경 직경(φ)',
        '대구경 공수 (n)',
        '대구경 환산직경(φd)',
        '공공과의 거리(a)',
        '장약장 (L)',
        '사각형의 변(W1)']

    for i in range(3, 15):
        ws[f'D{i}'] = sim1[i-3]

    sim2 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '공당 장약밀도(Ic)',
        '공당장약량(Q)',
        '적용장약량(Qa)',
        '장약장 (L)',
        '사각형의 변(W2)']

    for i in range(15, 23):
        ws[f'D{i}'] = sim2[i-15]

    sim3 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '공당 장약밀도(Ic)',
        '공당장약량(Q)',
        '적용장약량(Qa)',
        '장약장 (L)',
        '사각형의 변(W3)']

    for i in range(23, 31):
        ws[f'D{i}'] = sim3[i-23]

    sim4 = [
        '천공장 (H)',
        '저항선 (B)',
        '전색장 (h0)',
        '기저장 (hb)',
        '기저장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '공당장약량(Qc)',
        '적용 기저장약량(Qab)',
        '적용 주상장약량(Qac)',
        '총 장약량(Qtot)',
        '장약장 (L)',
        '사각형의 변(W4)']

    for i in range(31, 44):
        ws[f'D{i}'] = sim4[i-31]

    stoping = [
        '천공장 (H)',
        '저항선 (B)',
        '간격 (S)',
        '기저 장약장 (hb)',
        '전색장 (h0)',
        '기저 장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '주상장약량(Qc)',
        '적용기저장약량 (Qab)',
        '적용주상장약량 (Qac)',
        '총 공당장약량(Qtot)',
        '공수(n)',
        '장약장(L)']

    for i in range(47, 61):
        ws[f'D{i}'] = stoping[i-47]

    floor = [
        '천공장 (H)',
        '저항선 (B)',
        '간격 (S)',
        '기저 장약장 (hb)',
        '전색장 (h0)',
        '기저 장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '주상장약량(Qc)',
        '적용기저장약량 (Qab)',
        '적용주상장약량 (Qac)',
        '총 공당장약량(Qtot)',
        '공수(n)',
        '장약장(L)']

    for i in range(61, 75):
        ws[f'D{i}'] = floor[i-61]

    peri = [
        '천공장 (H)',
        '저항선 (B)',
        '간격 (S)',
        '기저 장약장 (hb)',
        '전색장 (h0)',
        '기저 장약밀도(Ib)',
        '주상 장약밀도(Ic)',
        '기저장약량(Qb)',
        '주상장약량(Qc)',
        '적용기저장약량 (Qab)',
        '적용주상장약량 (Qac)',
        '총 공당장약량(Qtot)',
        '공수(n)',
        '장약장(L)']

    for i in range(75, 89):
        ws[f'D{i}'] = peri[i-75]

    pcut, stop, flr, perim = PCutTest()

    ws['B1'].value = f'평행 심발공 패턴 설계 (H={pcut.cut1.H}m)'

    sim1 = [
        f'H = {pcut.cut1.H}m',
        f'B = {pcut.cut1.B : .3f}m',
        f'h0 = B = {pcut.cut1.h0: .3f}m',
        f'Ic = {pcut.cut1.Ic: .3f} kg/m',
        f'Qc = Ic * (H - h0) = {pcut.cut1.Ic: .3f} * ({pcut.cut1.H}-{pcut.cut1.h0: .3f}) = {pcut.cut1.Qc: .3f}kg',
        f'Qac = (Φ{pcut.cut1.IExp.phi}mm x {pcut.cut1.IExp.wei}kg x {pcut.cut1.IExp.len}mm) x {pcut.cut1.nc} = {pcut.cut1.Qca:.3f} kg ',
        f'φ = {pcut.cut1.Dia}m',
        f'n = {pcut.cut1.N}',
        f'φd = √{pcut.cut1.N} x {pcut.cut1.Dia}m = {math.sqrt(pcut.cut1.N)*pcut.cut1.Dia: .3f}',
        f'a = 1.5 x φd = {1.5 * math.sqrt(pcut.cut1.N)*pcut.cut1.Dia: .3f}',
        f'L = h0 + {pcut.cut1.IExp.len}mm x {pcut.cut1.nc}  = {pcut.cut1.L: .3f}',
        f'W1 = √2 x a = {pcut.cut1.W: .3f}']

    for i in range(3, 15):
        ws[f'F{i}'] = sim1[i-3]

    sim1 = [
        pcut.cut1.H,
        pcut.cut1.B,
        pcut.cut1.h0,
        pcut.cut1.Ic,
        pcut.cut1.Qc,
        pcut.cut1.Qca,
        pcut.cut1.Dia,
        pcut.cut1.N,
        math.sqrt(pcut.cut1.N)*pcut.cut1.Dia,
        1.5 * math.sqrt(pcut.cut1.N)*pcut.cut1.Dia,
        pcut.cut1.L,
        pcut.cut1.W]

    for i in range(3, 15):
        ws[f'G{i}'] = sim1[i-3]
        ws[f'G{i}'].number_format = '0.00'

    sim2 = [
        f'H = {pcut.cut2.H}m',
        f'B = {pcut.cut2.B : .3f}m',
        f'h0 = 0.5 x B = {pcut.cut2.h0: .3f}m',
        f'Ic = {pcut.cut2.Ic: .3f} kg/m',
        f'Qc = Ic * (H - h0) = {pcut.cut2.Ic: .3f} * ({pcut.cut2.H}-{pcut.cut2.h0: .3f}) = {pcut.cut2.Qc: .3f}kg',
        f'Qac = (Φ{pcut.cut2.IExp.phi}mm x {pcut.cut2.IExp.wei}kg x {pcut.cut2.IExp.len}mm) x {pcut.cut2.nc} = {pcut.cut2.Qca:.3f} kg ',
        f'L = h0 + {pcut.cut2.IExp.len}mm x {pcut.cut2.nc}  = {pcut.cut2.L: .3f}',
        f'W2 = √2 x a = {pcut.cut2.W: .3f}']

    for i in range(15, 23):
        ws[f'F{i}'] = sim2[i-15]

    sim2 = [
        pcut.cut2.H,
        pcut.cut2.B,
        pcut.cut2.h0,
        pcut.cut2.Ic,
        pcut.cut2.Qc,
        pcut.cut2.Qca,
        pcut.cut2.L,
        pcut.cut2.W]

    for i in range(15, 23):
        ws[f'G{i}'] = sim2[i-15]
        ws[f'G{i}'].number_format = '0.00'

    sim3 = [
        f'H = {pcut.cut3.H}m',
        f'B = {pcut.cut3.B : .3f}m',
        f'h0 = 0.5 x B = {pcut.cut3.h0: .3f}m',
        f'Ic = {pcut.cut3.Ic: .3f} kg/m',
        f'Qc = Ic * (H - h0) = {pcut.cut3.Ic: .3f} * ({pcut.cut3.H}-{pcut.cut3.h0: .3f}) = {pcut.cut3.Qc: .3f}kg',
        f'Qac = (Φ{pcut.cut3.IExp.phi}mm x {pcut.cut3.IExp.wei}kg x {pcut.cut3.IExp.len}mm) x {pcut.cut3.nc} = {pcut.cut3.Qca:.3f} kg ',
        f'L = h0 + {pcut.cut3.IExp.len}mm x {pcut.cut3.nc}  = {pcut.cut3.L: .3f}',
        f'W3 = √2 x a = {pcut.cut3.W: .3f}']

    for i in range(23, 31):
        ws[f'F{i}'] = sim3[i-23]

    sim3 = [
        pcut.cut3.H,
        pcut.cut3.B,
        pcut.cut3.h0,
        pcut.cut3.Ic,
        pcut.cut3.Qc,
        pcut.cut3.Qca,
        pcut.cut3.L,
        pcut.cut3.W]

    for i in range(23, 31):
        ws[f'G{i}'] = sim3[i-23]
        ws[f'G{i}'].number_format = '0.00'

    sim4 = [
        f'H = {pcut.cut4.H}m',
        f'B = {pcut.cut4.B : .3f}m',
        f'h0 = 0.5 x B =  {pcut.cut4.h0: .3f}m',
        f'hb = H/3 = {pcut.cut4.hb: .3f}m',
        f'Ib = {pcut.cut4.Ib: .3f} kg/m',
        f'Ic = {pcut.cut4.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {pcut.cut4.Ib: .3f} * {pcut.cut4.hb: .3f}) = {pcut.cut4.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {pcut.cut4.Ic: .3f} * ({pcut.cut4.H}-{pcut.cut4.h0: .3f}) = {pcut.cut4.Qc: .3f}kg',
        f'Qab = (Φ{pcut.cut4.IExp.phi}mm x {pcut.cut4.IExp.wei}kg x {pcut.cut4.IExp.len}mm) x {pcut.cut4.nb} = {pcut.cut4.Qba:.3f} kg ',
        f'Qac = (Φ{pcut.cut4.IExp.phi}mm x {pcut.cut4.IExp.wei}kg x {pcut.cut4.IExp.len}mm) x {pcut.cut4.nc} = {pcut.cut4.Qca:.3f} kg ',
        f'Qtot = {pcut.cut4.Qba:0.3f} + {pcut.cut4.Qca:0.3f} = {pcut.cut4.Qtot:.3f} kg ',
        f'L = h0 + {pcut.cut4.IExp.len}mm x {pcut.cut4.nc}  = {pcut.cut4.L: .3f}',
        f'W4 = √2 x a = {pcut.cut4.W: .3f}']

    for i in range(31, 44):
        ws[f'F{i}'] = sim4[i-31]

    sim4 = [
        pcut.cut4.H,
        pcut.cut4.B,
        pcut.cut4.h0,
        pcut.cut4.hb,
        pcut.cut4.Ib,
        pcut.cut4.Ic,
        pcut.cut4.Qb,
        pcut.cut4.Qc,
        pcut.cut4.Qba,
        pcut.cut4.Qca,
        pcut.cut4.Qtot,
        pcut.cut4.L,
        pcut.cut4.W]

    for i in range(31, 44):
        ws[f'G{i}'] = sim4[i-31]
        ws[f'G{i}'].number_format = '0.00'

    qtot = 4*(pcut.cut1.Qca + pcut.cut2.Qca + pcut.cut3.Qca + pcut.cut4.Qca + pcut.cut4.Qba)
    ws[f'G{44}'] = qtot
    ws[f'G{44}'].number_format = '0.00'
            
    vol = pcut.cut4.W * pcut.cut4.W * pcut.cut4.H
    ws[f'G{45}'] = vol
    ws[f'G{45}'].number_format = '0.00'

    ws[f'G{46}'] = qtot/vol
    ws[f'G{46}'].number_format = '0.00'

    stoping = [
        f'H = {stop.cut.H}m',
        f'B = {stop.cut.B : .3f}m',
        f'S = 1.1 x {stop.cut.B : .3f}m = {stop.cut.S}',
        f'hb = H/3 = {stop.cut.hb: .3f}m',
        f'h0 = 0.5 x B =  {stop.cut.h0: .3f}m',
        f'Ib = {stop.cut.Ib: .3f} kg/m',
        f'Ic = {stop.cut.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {stop.cut.Ib: .3f} * {stop.cut.hb: .3f}) = {stop.cut.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {stop.cut.Ic: .3f} * ({stop.cut.H}-{stop.cut.h0: .3f}) = {stop.cut.Qc: .3f}kg',
        f'Qab = (Φ{stop.cut.IExp.phi}mm x {stop.cut.IExp.wei}kg x {stop.cut.IExp.len}mm) x {stop.cut.nb} = {stop.cut.Qba:.3f} kg ',
        f'Qac = (Φ{stop.cut.IExp.phi}mm x {stop.cut.IExp.wei}kg x {stop.cut.IExp.len}mm) x {stop.cut.nc} = {stop.cut.Qca:.3f} kg ',
        f'Qtot = {stop.cut.Qba:0.3f} + {stop.cut.Qca:0.3f} = {stop.cut.Qtot:.3f} kg ',
        f'n = 1',
        f'L = h0 + {stop.cut.IExp.len}mm x {stop.cut.nc}  = {stop.cut.L: .3f}m', ]

    for i in range(47, 61):
        ws[f'F{i}'] = stoping[i-47]

    stoping = [
        stop.cut.H,
        stop.cut.B,
        stop.cut.B,
        stop.cut.hb,
        stop.cut.h0,
        stop.cut.Ib,
        stop.cut.Ic,
        stop.cut.Qb,
        stop.cut.Qc,
        stop.cut.Qba,
        stop.cut.Qca,
        stop.cut.Qtot,
        1,
        stop.cut.L]

    for i in range(47, 61):
        ws[f'G{i}'] = stoping[i-47]
        ws[f'G{i}'].number_format = '0.00'

    floor = [
        f'H = {flr.cut.H}m',
        f'B = {flr.cut.B : .3f}m',
        f'S = 1.1 x {flr.cut.B : .3f}m = {flr.cut.S}',
        f'hb = H/3 = {flr.cut.hb: .3f}m',
        f'h0 = 0.5 x B =  {flr.cut.h0: .3f}m',
        f'Ib = {flr.cut.Ib: .3f} kg/m',
        f'Ic = {flr.cut.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {flr.cut.Ib: .3f} * {flr.cut.hb: .3f}) = {flr.cut.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {flr.cut.Ic: .3f} * ({flr.cut.H}-{flr.cut.h0: .3f}) = {flr.cut.Qc: .3f}kg',
        f'Qab = (Φ{flr.cut.IExp.phi}mm x {flr.cut.IExp.wei}kg x {flr.cut.IExp.len}mm) x {flr.cut.nb} = {flr.cut.Qba:.3f} kg ',
        f'Qac = (Φ{flr.cut.IExp.phi}mm x {flr.cut.IExp.wei}kg x {flr.cut.IExp.len}mm) x {flr.cut.nc} = {flr.cut.Qca:.3f} kg ',
        f'Qtot = {flr.cut.Qba:0.3f} + {flr.cut.Qca:0.3f} = {flr.cut.Qtot:.3f} kg ',
        f'n = 1',
        f'L = h0 + {flr.cut.IExp.len}mm x {flr.cut.nc}  = {flr.cut.L: .3f}m', ]

    for i in range(61, 75):
        ws[f'F{i}'] = floor[i-61]

    floor = [
        flr.cut.H,
        flr.cut.B,
        flr.cut.S,
        flr.cut.hb,
        flr.cut.h0,
        flr.cut.Ib,
        flr.cut.Ic,
        flr.cut.Qb,
        flr.cut.Qc,
        flr.cut.Qba,
        flr.cut.Qca,
        flr.cut.Qtot,
        1,
        flr.cut.L]

    for i in range(61, 75):
        ws[f'G{i}'] = floor[i-61]
        ws[f'G{i}'].number_format = '0.00'

    peri = [
        f'H = {perim.cut.H}m',
        f'B = {perim.cut.B : .3f}m',
        f'S = 1.1 x {perim.cut.B : .3f}m = {perim.cut.S}',
        f'hb = H/3 = {perim.cut.hb: .3f}m',
        f'h0 = 0.5 x B =  {perim.cut.h0: .3f}m',
        f'Ib = {perim.cut.Ib: .3f} kg/m',
        f'Ic = {perim.cut.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {perim.cut.Ib: .3f} * {perim.cut.hb: .3f}) = {perim.cut.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {perim.cut.Ic: .3f} * ({perim.cut.H}-{perim.cut.h0: .3f}) = {perim.cut.Qc: .3f}kg',
        f'Qab = (Φ{perim.cut.IExp.phi}mm x {perim.cut.IExp.wei}kg x {perim.cut.IExp.len}mm) x {perim.cut.nb} = {perim.cut.Qba:.3f} kg ',
        f'Qac = (Φ{perim.cut.OExp.phi}mm x {perim.cut.OExp.wei}kg x {perim.cut.OExp.len}mm) x {perim.cut.nc} = {perim.cut.Qca:.3f} kg ',
        f'Qtot = {perim.cut.Qba:0.3f} + {perim.cut.Qca:0.3f} = {perim.cut.Qtot:.3f} kg ',
        f'n = 1',
        f'L = h0 + {perim.cut.IExp.len}mm x {perim.cut.nc}  = {perim.cut.L: .3f}m', ]

    for i in range(75, 89):
        ws[f'F{i}'] = peri[i-75]

    peri = [
        perim.cut.H,
        perim.cut.B,
        perim.cut.S,
        perim.cut.hb,
        perim.cut.h0,
        perim.cut.Ib,
        perim.cut.Ic,
        perim.cut.Qb,
        perim.cut.Qc,
        perim.cut.Qba,
        perim.cut.Qca,
        perim.cut.Qtot,
        1,
        perim.cut.L]

    for i in range(75, 89):
        ws[f'G{i}'] = peri[i-75]
        ws[f'G{i}'].number_format = '0.00'

    return wb


def xls_pr(wb):


    ws = wb.active

    pcut, stop, flr, perim = PCutTestPR()

    ws['B1'].value = f'평행 심발공 패턴 설계 (H={pcut.cut1.H}m)'

    sim1 = [
        f'H = {pcut.cut1.H}m',
        f'B = {pcut.cut1.B : .3f}m',
        f'h0 = B = {pcut.cut1.h0: .3f}m',
        f'Ic = {pcut.cut1.Ic: .3f} kg/m',
        f'Qc = Ic * (H - h0) = {pcut.cut1.Ic: .3f} * ({pcut.cut1.H}-{pcut.cut1.h0: .3f}) = {pcut.cut1.Qc: .3f}kg',
        f'Qac = (Φ{pcut.cut1.IExp.phi}mm x {pcut.cut1.IExp.wei}kg x {pcut.cut1.IExp.len}mm) x {pcut.cut1.nc} = {pcut.cut1.Qca:.3f} kg ',
        f'φ = {pcut.cut1.Dia}m',
        f'n = {pcut.cut1.N}',
        f'φd = √{pcut.cut1.N} x {pcut.cut1.Dia}m = {math.sqrt(pcut.cut1.N)*pcut.cut1.Dia: .3f}',
        f'a = 1.5 x φd = {1.5 * math.sqrt(pcut.cut1.N)*pcut.cut1.Dia: .3f}',
        f'L = h0 + {pcut.cut1.IExp.len}mm x {pcut.cut1.nc}  = {pcut.cut1.L: .3f}',
        f'W1 = √2 x a = {pcut.cut1.W: .3f}']

    for i in range(3, 15):
        ws[f'I{i}'] = sim1[i-3]

    sim1 = [
        pcut.cut1.H,
        pcut.cut1.B,
        pcut.cut1.h0,
        pcut.cut1.Ic,
        pcut.cut1.Qc,
        pcut.cut1.Qca,
        pcut.cut1.Dia,
        pcut.cut1.N,
        math.sqrt(pcut.cut1.N)*pcut.cut1.Dia,
        1.5 * math.sqrt(pcut.cut1.N)*pcut.cut1.Dia,
        pcut.cut1.L,
        pcut.cut1.W]

    for i in range(3, 15):
        ws[f'J{i}'] = sim1[i-3]
        ws[f'J{i}'].number_format = '0.00'

    sim2 = [
        f'H = {pcut.cut2.H}m',
        f'B = {pcut.cut2.B : .3f}m',
        f'h0 = 0.5 x B = {pcut.cut2.h0: .3f}m',
        f'Ic = {pcut.cut2.Ic: .3f} kg/m',
        f'Qc = Ic * (H - h0) = {pcut.cut2.Ic: .3f} * ({pcut.cut2.H}-{pcut.cut2.h0: .3f}) = {pcut.cut2.Qc: .3f}kg',
        f'Qac = (Φ{pcut.cut2.IExp.phi}mm x {pcut.cut2.IExp.wei}kg x {pcut.cut2.IExp.len}mm) x {pcut.cut2.nc} = {pcut.cut2.Qca:.3f} kg ',
        f'L = h0 + {pcut.cut2.IExp.len}mm x {pcut.cut2.nc}  = {pcut.cut2.L: .3f}',
        f'W2 = √2 x a = {pcut.cut2.W: .3f}']

    for i in range(15, 23):
        ws[f'I{i}'] = sim2[i-15]

    sim2 = [
        pcut.cut2.H,
        pcut.cut2.B,
        pcut.cut2.h0,
        pcut.cut2.Ic,
        pcut.cut2.Qc,
        pcut.cut2.Qca,
        pcut.cut2.L,
        pcut.cut2.W]

    for i in range(15, 23):
        ws[f'J{i}'] = sim2[i-15]
        ws[f'J{i}'].number_format = '0.00'

    sim3 = [
        f'H = {pcut.cut3.H}m',
        f'B = {pcut.cut3.B : .3f}m',
        f'h0 = 0.5 x B = {pcut.cut3.h0: .3f}m',
        f'Ic = {pcut.cut3.Ic: .3f} kg/m',
        f'Qc = Ic * (H - h0) = {pcut.cut3.Ic: .3f} * ({pcut.cut3.H}-{pcut.cut3.h0: .3f}) = {pcut.cut3.Qc: .3f}kg',
        f'Qac = (Φ{pcut.cut3.IExp.phi}mm x {pcut.cut3.IExp.wei}kg x {pcut.cut3.IExp.len}mm) x {pcut.cut3.nc} = {pcut.cut3.Qca:.3f} kg ',
        f'L = h0 + {pcut.cut3.IExp.len}mm x {pcut.cut3.nc}  = {pcut.cut3.L: .3f}',
        f'W3 = √2 x a = {pcut.cut3.W: .3f}']

    for i in range(23, 31):
        ws[f'I{i}'] = sim3[i-23]

    sim3 = [
        pcut.cut3.H,
        pcut.cut3.B,
        pcut.cut3.h0,
        pcut.cut3.Ic,
        pcut.cut3.Qc,
        pcut.cut3.Qca,
        pcut.cut3.L,
        pcut.cut3.W]

    for i in range(23, 31):
        ws[f'J{i}'] = sim3[i-23]
        ws[f'J{i}'].number_format = '0.00'

    sim4 = [
        f'H = {pcut.cut4.H}m',
        f'B = {pcut.cut4.B : .3f}m',
        f'h0 = 0.5 x B =  {pcut.cut4.h0: .3f}m',
        f'hb = H/3 = {pcut.cut4.hb: .3f}m',
        f'Ib = {pcut.cut4.Ib: .3f} kg/m',
        f'Ic = {pcut.cut4.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {pcut.cut4.Ib: .3f} * {pcut.cut4.hb: .3f}) = {pcut.cut4.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {pcut.cut4.Ic: .3f} * ({pcut.cut4.H}-{pcut.cut4.h0: .3f}) = {pcut.cut4.Qc: .3f}kg',
        f'Qab = (Φ{pcut.cut4.IExp.phi}mm x {pcut.cut4.IExp.wei}kg x {pcut.cut4.IExp.len}mm) x {pcut.cut4.nb} = {pcut.cut4.Qba:.3f} kg ',
        f'Qac = (Φ{pcut.cut4.IExp.phi}mm x {pcut.cut4.IExp.wei}kg x {pcut.cut4.IExp.len}mm) x {pcut.cut4.nc} = {pcut.cut4.Qca:.3f} kg ',
        f'Qtot = {pcut.cut4.Qba:0.3f} + {pcut.cut4.Qca:0.3f} = {pcut.cut4.Qtot:.3f} kg ',
        f'L = h0 + {pcut.cut4.IExp.len}mm x {pcut.cut4.nc}  = {pcut.cut4.L: .3f}',
        f'W4 = √2 x a = {pcut.cut4.W: .3f}']

    for i in range(31, 44):
        ws[f'I{i}'] = sim4[i-31]

    sim4 = [
        pcut.cut4.H,
        pcut.cut4.B,
        pcut.cut4.h0,
        pcut.cut4.hb,
        pcut.cut4.Ib,
        pcut.cut4.Ic,
        pcut.cut4.Qb,
        pcut.cut4.Qc,
        pcut.cut4.Qba,
        pcut.cut4.Qca,
        pcut.cut4.Qtot,
        pcut.cut4.L,
        pcut.cut4.W]

    for i in range(31, 44):
        ws[f'J{i}'] = sim4[i-31]
        ws[f'J{i}'].number_format = '0.00'

    stoping = [
        f'H = {stop.cut.H}m',
        f'B = {stop.cut.B : .3f}m',
        f'S = 1.1 x {stop.cut.B : .3f}m = {stop.cut.S}',
        f'hb = H/3 = {stop.cut.hb: .3f}m',
        f'h0 = 0.5 x B =  {stop.cut.h0: .3f}m',
        f'Ib = {stop.cut.Ib: .3f} kg/m',
        f'Ic = {stop.cut.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {stop.cut.Ib: .3f} * {stop.cut.hb: .3f}) = {stop.cut.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {stop.cut.Ic: .3f} * ({stop.cut.H}-{stop.cut.h0: .3f}) = {stop.cut.Qc: .3f}kg',
        f'Qab = (Φ{stop.cut.IExp.phi}mm x {stop.cut.IExp.wei}kg x {stop.cut.IExp.len}mm) x {stop.cut.nb} = {stop.cut.Qba:.3f} kg ',
        f'Qac = (Φ{stop.cut.IExp.phi}mm x {stop.cut.IExp.wei}kg x {stop.cut.IExp.len}mm) x {stop.cut.nc} = {stop.cut.Qca:.3f} kg ',
        f'Qtot = {stop.cut.Qba:0.3f} + {stop.cut.Qca:0.3f} = {stop.cut.Qtot:.3f} kg ',
        f'n = 1',
        f'L = h0 + {stop.cut.IExp.len}mm x {stop.cut.nc}  = {stop.cut.L: .3f}m', ]

    for i in range(47, 61):
        ws[f'I{i}'] = stoping[i-47]

    stoping = [
        stop.cut.H,
        stop.cut.B,
        stop.cut.B,
        stop.cut.hb,
        stop.cut.h0,
        stop.cut.Ib,
        stop.cut.Ic,
        stop.cut.Qb,
        stop.cut.Qc,
        stop.cut.Qba,
        stop.cut.Qca,
        stop.cut.Qtot,
        1,
        stop.cut.L]

    for i in range(47, 61):
        ws[f'J{i}'] = stoping[i-47]
        ws[f'J{i}'].number_format = '0.00'

    floor = [
        f'H = {flr.cut.H}m',
        f'B = {flr.cut.B : .3f}m',
        f'S = 1.1 x {flr.cut.B : .3f}m = {flr.cut.S}',
        f'hb = H/3 = {flr.cut.hb: .3f}m',
        f'h0 = 0.5 x B =  {flr.cut.h0: .3f}m',
        f'Ib = {flr.cut.Ib: .3f} kg/m',
        f'Ic = {flr.cut.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {flr.cut.Ib: .3f} * {flr.cut.hb: .3f}) = {flr.cut.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {flr.cut.Ic: .3f} * ({flr.cut.H}-{flr.cut.h0: .3f}) = {flr.cut.Qc: .3f}kg',
        f'Qab = (Φ{flr.cut.IExp.phi}mm x {flr.cut.IExp.wei}kg x {flr.cut.IExp.len}mm) x {flr.cut.nb} = {flr.cut.Qba:.3f} kg ',
        f'Qac = (Φ{flr.cut.IExp.phi}mm x {flr.cut.IExp.wei}kg x {flr.cut.IExp.len}mm) x {flr.cut.nc} = {flr.cut.Qca:.3f} kg ',
        f'Qtot = {flr.cut.Qba:0.3f} + {flr.cut.Qca:0.3f} = {flr.cut.Qtot:.3f} kg ',
        f'n = 1',
        f'L = h0 + {flr.cut.IExp.len}mm x {flr.cut.nc}  = {flr.cut.L: .3f}m', ]

    for i in range(61, 75):
        ws[f'I{i}'] = floor[i-61]

    floor = [
        flr.cut.H,
        flr.cut.B,
        flr.cut.S,
        flr.cut.hb,
        flr.cut.h0,
        flr.cut.Ib,
        flr.cut.Ic,
        flr.cut.Qb,
        flr.cut.Qc,
        flr.cut.Qba,
        flr.cut.Qca,
        flr.cut.Qtot,
        1,
        flr.cut.L]

    for i in range(61, 75):
        ws[f'J{i}'] = floor[i-61]
        ws[f'J{i}'].number_format = '0.00'

    peri = [
        f'H = {perim.cut.H}m',
        f'B = {perim.cut.B : .3f}m',
        f'S = 1.1 x {perim.cut.B : .3f}m = {perim.cut.S}',
        f'hb = H/3 = {perim.cut.hb: .3f}m',
        f'h0 = 0.5 x B =  {perim.cut.h0: .3f}m',
        f'Ib = {perim.cut.Ib: .3f} kg/m',
        f'Ic = {perim.cut.Ic: .3f} kg/m',
        f'Qb = Ib * hb = {perim.cut.Ib: .3f} * {perim.cut.hb: .3f}) = {perim.cut.Qb: .3f}kg',
        f'Qc = Ic * (H - h0) = {perim.cut.Ic: .3f} * ({perim.cut.H}-{perim.cut.h0: .3f}) = {perim.cut.Qc: .3f}kg',
        f'Qab = (Φ{perim.cut.IExp.phi}mm x {perim.cut.IExp.wei}kg x {perim.cut.IExp.len}mm) x {perim.cut.nb} = {perim.cut.Qba:.3f} kg ',
        f'Qac = (Φ{perim.cut.OExp.phi}mm x {perim.cut.OExp.wei}kg x {perim.cut.OExp.len}mm) x {perim.cut.nc} = {perim.cut.Qca:.3f} kg ',
        f'Qtot = {perim.cut.Qba:0.3f} + {perim.cut.Qca:0.3f} = {perim.cut.Qtot:.3f} kg ',
        f'n = 1',
        f'L = h0 + {perim.cut.IExp.len}mm x {perim.cut.nc}  = {perim.cut.L: .3f}m', ]

    for i in range(75, 89):
        ws[f'I{i}'] = peri[i-75]

    peri = [
        perim.cut.H,
        perim.cut.B,
        perim.cut.S,
        perim.cut.hb,
        perim.cut.h0,
        perim.cut.Ib,
        perim.cut.Ic,
        perim.cut.Qb,
        perim.cut.Qc,
        perim.cut.Qba,
        perim.cut.Qca,
        perim.cut.Qtot,
        1,
        perim.cut.L]

    for i in range(75, 89):
        ws[f'J{i}'] = peri[i-75]
        ws[f'J{i}'].number_format = '0.00'

    return wb


def main():
    wb = Workbook()    
    # rnd_test()
    # vcut()
    wb = form_bl(wb)    
    wb = xls_tb(wb)
    wb = xls_pr(wb)
    wb.save("d:/tset.xlsx")

if __name__ == "__main__":
    main()
